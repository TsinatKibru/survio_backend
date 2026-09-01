from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import authentication, permissions, status as drf_status

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from forms_builder.models import Form, ReportingPeriod
from accounts.permissions import IsAdminOrAbove
from submissions.models import Answer

from .services import FormReportBuilder, PDFReportRenderer


# ─── Shared helpers ────────────────────────────────────────────────────────────

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'),
)

HEADER_FILL = PatternFill(start_color='2E5FA3', end_color='2E5FA3', fill_type='solid')
ALT_FILL = PatternFill(start_color='EDF2FB', end_color='EDF2FB', fill_type='solid')
SECTION_FILL = PatternFill(start_color='D6E4F7', end_color='D6E4F7', fill_type='solid')


# ─── Views ─────────────────────────────────────────────────────────────────────

class FormReportListView(APIView):
    """
    Lists available forms for report generation, with their reporting periods.

    GET /api/reports/forms/
    """
    authentication_classes = [
        authentication.SessionAuthentication,
        authentication.BasicAuthentication,
    ]
    permission_classes = [IsAdminOrAbove]

    def get(self, request):
        forms = Form.objects.filter(is_active=True).prefetch_related('periods', 'category')
        data = []
        for form in forms:
            periods = [
                {
                    'id': p.id,
                    'label': p.label,
                    'period_start': str(p.period_start),
                    'period_end': str(p.period_end),
                    'status': p.status,
                }
                for p in form.periods.all()
            ]
            data.append({
                'form_id': form.id,
                'title': form.title,
                'category': form.category.name if form.category else None,
                'schedule_type': form.schedule_type,
                'periods': periods,
                'report_urls': {
                    'summary': request.build_absolute_uri(f'/api/reports/forms/{form.id}/summary/'),
                    'pdf': request.build_absolute_uri(f'/api/reports/forms/{form.id}/pdf/'),
                    'excel': request.build_absolute_uri(f'/api/reports/forms/{form.id}/excel/'),
                },
            })
        return Response(data)


class FormReportSummaryView(APIView):
    """
    Returns full JSON analytics for a form.

    GET /api/reports/forms/<form_id>/summary/
    Query params (all optional):
      ?period_id=<int>      — filter to one ReportingPeriod
      ?industry_id=<int>    — filter to one factory
      ?category_id=<int>    — filter to one food category
    """
    authentication_classes = [
        authentication.SessionAuthentication,
        authentication.BasicAuthentication,
    ]
    permission_classes = [IsAdminOrAbove]

    def get(self, request, form_id):
        period_id = request.query_params.get('period_id')
        industry_id = request.query_params.get('industry_id')
        category_id = request.query_params.get('category_id')

        try:
            builder = FormReportBuilder()
            report = builder.build(
                form_id=form_id,
                period_id=int(period_id) if period_id else None,
                industry_id=int(industry_id) if industry_id else None,
                category_id=int(category_id) if category_id else None,
            )
        except Form.DoesNotExist:
            return Response({'error': f'Form {form_id} not found.'}, status=drf_status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({'error': str(e)}, status=drf_status.HTTP_500_INTERNAL_SERVER_ERROR)

        return Response(report)


class FormReportPDFView(APIView):
    """
    Generates and downloads a KoboToolbox-style PDF report for a form.

    GET /api/reports/forms/<form_id>/pdf/
    Query params: same as FormReportSummaryView
    """
    authentication_classes = [
        authentication.SessionAuthentication,
        authentication.BasicAuthentication,
    ]
    permission_classes = [IsAdminOrAbove]

    def get(self, request, form_id):
        period_id = request.query_params.get('period_id')
        industry_id = request.query_params.get('industry_id')
        category_id = request.query_params.get('category_id')

        try:
            builder = FormReportBuilder()
            report = builder.build(
                form_id=form_id,
                period_id=int(period_id) if period_id else None,
                industry_id=int(industry_id) if industry_id else None,
                category_id=int(category_id) if category_id else None,
            )
        except Form.DoesNotExist:
            return Response({'error': f'Form {form_id} not found.'}, status=drf_status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({'error': str(e)}, status=drf_status.HTTP_500_INTERNAL_SERVER_ERROR)

        try:
            renderer = PDFReportRenderer()
            pdf_buffer = renderer.render(report)
        except Exception as e:
            return Response({'error': f'PDF generation failed: {str(e)}'}, status=drf_status.HTTP_500_INTERNAL_SERVER_ERROR)

        safe_title = ''.join(c if c.isalnum() or c in ' _-' else '_' for c in report['form_title'])
        filename = f"report_{safe_title[:40]}_{report['generated_at'][:10]}.pdf"

        response = HttpResponse(pdf_buffer.read(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


class FormReportExcelView(APIView):
    """
    Generates and downloads an Excel workbook for a form.
    Tab 1: Aggregated Summary (all questions, stats, frequencies).
    Tab 2: Raw Submissions Matrix (rows = submissions, cols = questions).

    GET /api/reports/forms/<form_id>/excel/
    Query params: same as FormReportSummaryView
    """
    authentication_classes = [
        authentication.SessionAuthentication,
        authentication.BasicAuthentication,
    ]
    permission_classes = [IsAdminOrAbove]

    def get(self, request, form_id):
        period_id = request.query_params.get('period_id')
        industry_id = request.query_params.get('industry_id')
        category_id = request.query_params.get('category_id')

        try:
            builder = FormReportBuilder()
            report = builder.build(
                form_id=form_id,
                period_id=int(period_id) if period_id else None,
                industry_id=int(industry_id) if industry_id else None,
                category_id=int(category_id) if category_id else None,
            )
        except Form.DoesNotExist:
            return Response({'error': f'Form {form_id} not found.'}, status=drf_status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({'error': str(e)}, status=drf_status.HTTP_500_INTERNAL_SERVER_ERROR)

        wb = self._build_workbook(report, form_id)

        import io
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        safe_title = ''.join(c if c.isalnum() or c in ' _-' else '_' for c in report['form_title'])
        filename = f"report_{safe_title[:40]}_{report['generated_at'][:10]}.xlsx"

        response = HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    def _build_workbook(self, report, form_id):
        wb = Workbook()

        # ── Tab 1: Aggregated Summary ──────────────────────────────────────
        ws1 = wb.active
        ws1.title = 'Summary'

        # Title row
        ws1.append([report['form_title']])
        ws1['A1'].font = Font(bold=True, size=14, color='FFFFFF')
        ws1['A1'].fill = HEADER_FILL
        ws1.merge_cells('A1:F1')
        ws1.append([
            f"Generated: {report['generated_at'][:19].replace('T', ' ')}",
            f"Submissions: {report['total_submissions']}",
            f"Filters: {', '.join(report['filters_applied'])}",
        ])
        ws1.append([])  # blank row

        TYPE_LABELS = {
            'select': 'SELECT_ONE', 'multiselect': 'SELECT_MULTIPLE',
            'number': 'INTEGER', 'decimal': 'DECIMAL',
            'text': 'TEXT', 'textarea': 'TEXT', 'yes_no': 'SELECT_ONE',
            'image': 'IMAGE', 'phone': 'TEXT', 'email': 'TEXT',
            'location': 'LOCATION', 'date': 'DATE',
        }

        for section in report['sections']:
            # Section heading row
            ws1.append([section['title']])
            sec_row = ws1.max_row
            ws1.merge_cells(f'A{sec_row}:F{sec_row}')
            ws1[f'A{sec_row}'].fill = SECTION_FILL
            ws1[f'A{sec_row}'].font = Font(bold=True, size=11)
            ws1.append([])

            for q in section['questions']:
                type_badge = TYPE_LABELS.get(q['type'], q['type'].upper())
                ws1.append([q['label'], f"TYPE: {type_badge}",
                             f"Answered: {q['answered_count']}/{q['total_count']}"])
                label_row = ws1.max_row
                ws1[f'A{label_row}'].font = Font(bold=True)

                if 'stats' in q:
                    s = q['stats']
                    ws1.append(['Mean', 'Median', 'Mode', 'Std Dev', 'Min', 'Max', 'Sum'])
                    hrow = ws1.max_row
                    for col_idx in range(1, 8):
                        cell = ws1.cell(row=hrow, column=col_idx)
                        cell.fill = HEADER_FILL
                        cell.font = Font(bold=True, color='FFFFFF')
                        cell.alignment = Alignment(horizontal='center')
                    ws1.append([s['mean'], s['median'], s['mode'],
                                 s['std_dev'], s['min'], s['max'], s['sum']])
                    for col_idx in range(1, 8):
                        ws1.cell(row=ws1.max_row, column=col_idx).alignment = Alignment(horizontal='center')

                if 'choices' in q and q['choices']:
                    ws1.append(['Option / Value', 'Count', 'Percentage'])
                    hrow = ws1.max_row
                    for col_idx, col_name in enumerate(['A', 'B', 'C'], 1):
                        cell = ws1.cell(row=hrow, column=col_idx)
                        cell.fill = HEADER_FILL
                        cell.font = Font(bold=True, color='FFFFFF')
                    for i, row in enumerate(q['choices']):
                        ws1.append([row['label'], row['count'], f"{row['percentage']}%"])
                        if i % 2 == 1:
                            for col_idx in range(1, 4):
                                ws1.cell(row=ws1.max_row, column=col_idx).fill = ALT_FILL

                if 'text_responses' in q and q['text_responses']:
                    ws1.append(['Text Response', 'Count'])
                    hrow = ws1.max_row
                    for col_idx in range(1, 3):
                        ws1.cell(row=hrow, column=col_idx).fill = HEADER_FILL
                        ws1.cell(row=hrow, column=col_idx).font = Font(bold=True, color='FFFFFF')
                    for r in q['text_responses']:
                        ws1.append([r['value'], r['count']])

                if 'upload_count' in q:
                    ws1.append([f"Total media uploads: {q['upload_count']}"])

                ws1.append([])  # spacing

        # Auto-width col A
        ws1.column_dimensions['A'].width = 55
        for col in ['B', 'C', 'D', 'E', 'F', 'G']:
            ws1.column_dimensions[col].width = 14

        # ── Tab 2: Raw Submissions Matrix ──────────────────────────────────
        ws2 = wb.create_sheet('Raw Data')
        self._build_raw_sheet(ws2, report, form_id)

        return wb

    def _build_raw_sheet(self, ws, report, form_id):
        """Builds the raw submission matrix: rows = submissions, cols = questions."""
        from submissions.models import Submission, Answer
        from django.db.models import Q

        # Collect all questions in order
        all_questions = []
        for section in report['sections']:
            for q_data in section['questions']:
                all_questions.append(q_data)

        if not all_questions:
            ws.append(['No questions found.'])
            return

        q_ids = [q['question_id'] for q in all_questions]

        # Header row: metadata + question labels
        headers = ['Submission ID', 'Organization', 'Submitted At', 'Period', 'On Time']
        headers += [f"Q{i+1}: {q['label'][:40]}" for i, q in enumerate(all_questions)]
        ws.append(headers)

        hrow = ws.max_row
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=hrow, column=col_idx)
            cell.fill = HEADER_FILL
            cell.font = Font(bold=True, color='FFFFFF', size=9)
            cell.alignment = Alignment(horizontal='center', wrap_text=True)

        # Fetch submissions for this form
        try:
            form = Form.objects.get(pk=form_id)
        except Form.DoesNotExist:
            ws.append(['Form not found.'])
            return

        submissions = Submission.objects.filter(
            form=form, status=Submission.STATUS_SUBMITTED
        ).select_related('organization', 'period').order_by('-submitted_at')

        for i, sub in enumerate(submissions):
            # Prefetch all answers for this submission
            answers_map = {
                a.question_id: a.value
                for a in Answer.objects.filter(submission=sub, question_id__in=q_ids)
            }

            row = [
                sub.id,
                sub.organization.name if sub.organization else sub.industry_name,
                sub.submitted_at.strftime('%Y-%m-%d %H:%M') if sub.submitted_at else '',
                sub.period.label if sub.period else '',
                'Yes' if not sub.is_late else 'No',
            ]
            row += [answers_map.get(q['question_id'], '') for q in all_questions]
            ws.append(row)

            if i % 2 == 1:
                for col_idx in range(1, len(headers) + 1):
                    ws.cell(row=ws.max_row, column=col_idx).fill = ALT_FILL

        # Column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 16
        ws.column_dimensions['E'].width = 10
        for i in range(6, len(headers) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 22
